import csv
from pathlib import Path

from openpyxl import load_workbook


def try_count_rows(path: Path) -> int | None:
    extension = path.suffix.lower().lstrip(".")
    try:
        if extension in {"csv", "txt"}:
            with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as handle:
                return max(sum(1 for _ in csv.reader(handle)) - 1, 0)
        if extension == "xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            worksheet = workbook.active
            return max((worksheet.max_row or 0) - 1, 0)
    except Exception:
        return None
    return None

